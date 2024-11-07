# -*- coding: utf-8 -*-

# 20240715:引入EKF融合速度
# 20240716:EKF融合速度计算完后增加滑动窗口平移
# 20240720:8 radar_0.3 9 radar_0.5
import serial
import serial.tools.list_ports
import sys
import random
import math
from datetime import datetime
from openpyxl import Workbook
import re
import socket
import base64
import threading
from PyQt5 import QtCore
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from WelcomeScreen_314 import WelcomeScreen
from cesuUI import cesu_UI
from pynmea_fun import *
from EKF_Fixed_QR_and_Sliding_Window_user import *

# 定义全局变量 radar_speed
radar_speed = 0.0
acc_y = 0.0
acc_x = 0.0
yaw = 0.0


class WelcomeWindow(QWidget):
    def __init__(self, stacked_widget):
        super().__init__()
        self.stacked_widget = stacked_widget
        self.ui = WelcomeScreen()
        self.ui.setupUi(self)
        self.setFixedSize(self.size())
        self.ui.pushButton.clicked.connect(self.switch_window)

    def switch_window(self):
        self.stacked_widget.setCurrentIndex(1)


def handle_collected_gnss_data(speed_data, radar_data, acc_x_data, acc_y_data, yaw_data, estimated_Speed_data,
                               time_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Speed Data"
    ws.append(["Time Stamp", "Radar Data", "Acc X Data", "Acc Y Data", "Yaw", "Speed", "Estimated Speed"])

    for time_stamp, radar, acc_x, acc_y, yaw, speed, estimated_speed in zip(time_data, radar_data, acc_x_data,
                                                                            acc_y_data, yaw_data, speed_data,
                                                                            estimated_Speed_data):
        ws.append([time_stamp.strftime("%Y-%m-%d %H:%M:%S"), radar, acc_x, acc_y, yaw, speed, estimated_speed])

    file_name = datetime.now().strftime("%Y-%m-%d_%H-%M-%S") + "_gnss_speed.xlsx"
    wb.save(file_name)
    print(f"Data saved to {file_name}")


def handle_collected_wheel_data(speed_data, time_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Speed Data"
    ws.append(["Time Stamp", "Speed"])
    for time_stamp, speed in zip(time_data, speed_data):
        ws.append([time_stamp.strftime("%Y-%m-%d %H:%M:%S"), speed])
    file_name = datetime.now().strftime("%Y-%m-%d_%H-%M-%S") + "_wheel_speed.xlsx"
    wb.save(file_name)
    print(f"Data saved to {file_name}")


def handle_collected_radar_data(speed_data, time_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Radar Data"
    ws.append(["Time Stamp", "Speed"])
    for time_stamp, speed in zip(time_data, speed_data):
        ws.append([time_stamp.strftime("%Y-%m-%d %H:%M:%S"), speed])
    file_name = datetime.now().strftime("%Y-%m-%d_%H-%M-%S") + "_radar_data.xlsx"
    wb.save(file_name)
    print(f"Data saved to {file_name}")
    pass


def handle_collected_imu_data(acc_x, acc_y, yaw, time_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "IMU Data"
    ws.append(["Time Stamp", "acc_x", "acc_y", "yaw"])

    for time_stamp, ax, ay, y in zip(time_data, acc_x, acc_y, yaw):
        ws.append([time_stamp.strftime("%Y-%m-%d %H:%M:%S"), ax, ay, y])

    file_name = datetime.now().strftime("%Y-%m-%d_%H-%M-%S") + "_imu_data.xlsx"
    wb.save(file_name)
    print(f"Data saved to {file_name}")


class information_show_Window(QWidget):
    def __init__(self, stacked_widget):
        super().__init__()
        self.gnss_work = None
        self.radar_work = None
        self.wheel_work = None
        self.imu_work = None
        self.ui = cesu_UI()
        self.ui.setupUi(self)
        self.ui_process()
        self.port = None
        self.baudrate = None
        self.stacked_widget = stacked_widget
        self.speed_data_gnss = []
        self.speed_data_wheel = []
        self.speed_data_radar = []
        self.imu_data = []

    def ui_process(self):
        self.speed_show_lcd_for_wheel = self.ui.lcdNumber
        self.speed_show_lcd_for_radar = self.ui.lcdNumber_2
        self.speed_show_lcd_for_GNSS = self.ui.lcdNumber_3
        self.X_acc_show_lcd = self.ui.lcdNumber_4
        self.Y_acc_show_lcd = self.ui.lcdNumber_5
        self.yaw_show_lcd = self.ui.lcdNumber_6
        self.estimated_speed_show_lcd = self.ui.lcdNumber_7

        self.open_speed_show_for_wheel = self.ui.pushButton
        self.close_speed_show_for_wheel = self.ui.pushButton_2

        self.open_speed_show_for_radar = self.ui.pushButton_3
        self.close_speed_show_for_radar = self.ui.pushButton_4

        self.open_speed_show_for_gnss = self.ui.pushButton_5
        self.close_speed_show_for_gnss = self.ui.pushButton_6

        self.open_speed_show_for_IMU = self.ui.pushButton_7
        self.close_speed_show_for_IMU = self.ui.pushButton_8

        self.open_speed_show_for_wheel.clicked.connect(self.wheel_choose_port)
        self.close_speed_show_for_wheel.clicked.connect(self.wheel_speed_fun_end)

        self.open_speed_show_for_radar.clicked.connect(self.radar_choose_port)
        self.close_speed_show_for_radar.clicked.connect(self.radar_speed_fun_end)

        self.open_speed_show_for_gnss.clicked.connect(self.gnss_choose_port)
        self.open_speed_show_for_gnss.clicked.connect(self.ntrip_start)
        self.close_speed_show_for_gnss.clicked.connect(self.gnss_speed_fun_end)
        self.close_speed_show_for_gnss.clicked.connect(self.ntrip_stop)

        self.open_speed_show_for_IMU.clicked.connect(self.IMU_choose_port)
        self.close_speed_show_for_IMU.clicked.connect(self.imu_speed_fun_end)

    def wheel_choose_port(self):
        self.dialog_wheel = PortSelectionDialog_wheel()
        self.dialog_wheel.port_baudrate_selected_signal.connect(self.whell_port_baudrate_selected)
        self.dialog_wheel.exec_()

    def whell_port_baudrate_selected(self, chosen_port, chosen_baudrate):
        try:
            self.now = datetime.now()
            self.wheel_work = Worker_wheel(chosen_port, chosen_baudrate)
            self.wheel_work.finished_signal.connect(self.wheel_worker_finished)
            self.wheel_work.data_collected_signal.connect(handle_collected_wheel_data)
            self.wheel_work.tran_speed_signal.connect(self.speed_wheel_show)
            self.wheel_work.mistake_message_transmit.connect(self.mistake_message_show)
            self.wheel_work.start()
        except serial.serialutil.SerialException as error:
            self.mistake_message_show(str(error))

    def wheel_speed_fun_end(self):
        if self.wheel_work is not None:
            self.wheel_work.should_run = False
            self.wheel_work.quit()
            self.wheel_work.wait()

    def wheel_worker_finished(self):
        self.wheel_work = None

    def speed_wheel_show(self, speed):
        self.speed_show_lcd_for_wheel.display(speed)
        self.speed_show_lcd_for_wheel.show()

    def radar_choose_port(self):
        self.dialog_radar = PortSelectionDialog_radar()
        self.dialog_radar.port_baudrate_selected_signal.connect(self.radar_port_baudrate_selected)
        self.dialog_radar.exec_()

    def radar_port_baudrate_selected(self, chosen_port, chosen_baudrate):
        try:
            self.radar_now = datetime.now()
            self.radar_work = Worker_radar(chosen_port, chosen_baudrate)
            self.radar_work.finished_signal.connect(self.radar_worker_finished)
            self.radar_work.data_collected_signal.connect(handle_collected_radar_data)
            self.radar_work.avg_data_parsed.connect(self.radar_data_show)
            self.radar_work.mistake_message_transmit.connect(self.mistake_message_show)
            self.radar_work.start()
        except serial.serialutil.SerialException as error:
            self.mistake_message_show(error)

    def radar_speed_fun_end(self):
        if self.radar_work is not None:
            self.radar_work.should_run = False
            self.radar_work.quit()
            self.radar_work.wait()

    def radar_worker_finished(self):
        self.radar_work = None

    def radar_data_show(self, speed):
        self.speed_show_lcd_for_radar.display(speed)
        self.speed_show_lcd_for_radar.show()

    def gnss_choose_port(self):
        try:
            self.gnss_work = Worker_gnss()
            self.gnss_work.finished_signal.connect(self.gnss_worker_finished)
            self.gnss_work.average_speed_signal.connect(self.speed_gnss_show)
            self.gnss_work.estimated_speed_signal.connect(self.estimated_speed_gnss_show)
            self.gnss_work.data_collected_signal.connect(handle_collected_gnss_data)
            self.gnss_work.mistake_message_transmit.connect(self.mistake_message_show)
            self.gnss_work.start()
        except serial.serialutil.SerialException as error:
            self.mistake_message_show(error)

    def gnss_speed_fun_end(self):
        if self.gnss_work is not None:
            self.gnss_work.should_run = False
            self.gnss_work.quit()
            self.gnss_work.wait()

    def ntrip_start(self):
        try:
            self.ntrip_work = Worker_Ntrip(self.gnss_work)
            self.ntrip_work.finished_signal.connect(self.ntrip_worker_finished)
            self.ntrip_work.mistake_message_transmit.connect(self.mistake_message_show)
            self.ntrip_work.start()
        except serial.serialutil.SerialException as error:
            self.mistake_message_show(error)

    def ntrip_stop(self):
        if hasattr(self, 'ntrip_work') and self.gnss_work is not None:
            self.ntrip_work.should_run = False
            self.ntrip_work.quit()
            self.ntrip_work.wait()

    def ntrip_worker_finished(self):
        self.ntrip_work = None

    def gnss_worker_finished(self):
        self.gnss_work = None

    def speed_gnss_show(self, speed):
        self.speed_show_lcd_for_GNSS.display(speed)
        self.speed_show_lcd_for_GNSS.show()
        pass

    def estimated_speed_gnss_show(self, speed):
        self.estimated_speed_show_lcd.display(speed)
        self.estimated_speed_show_lcd.show()

    def IMU_choose_port(self):
        self.dialog_imu = PortSelectionDialog_imu()
        self.dialog_imu.port_baudrate_selected_signal.connect(self.imu_port_baudrate_selected)
        self.dialog_imu.exec_()

    def imu_port_baudrate_selected(self, chosen_port, chosen_baudrate):
        try:
            self.now = datetime.now()
            self.imu_work = Worker_IMU(chosen_port, chosen_baudrate)
            self.imu_work.finished_signal.connect(self.imu_worker_finished)
            self.imu_work.tran_data_collected_signal.connect(handle_collected_imu_data)
            self.imu_work.tran_x_acc_signal.connect(self.x_acc_show)
            self.imu_work.tran_y_acc_signal.connect(self.y_acc_show)
            self.imu_work.tran_yaw_signal.connect(self.yaw_show)
            self.imu_work.mistake_message_transmit.connect(self.mistake_message_show)
            self.imu_work.start()
        except serial.serialutil.SerialException as error:
            self.mistake_message_show(str(error))

    def imu_speed_fun_end(self):
        if self.imu_work is not None:
            try:
                self.imu_work.should_run = False
                self.imu_work.quit()
                self.imu_work.wait()
            except Exception as e:
                self.mistake_message_show(f"Error stopping IMU work: {e}")
            finally:
                self.imu_worker_finished()

    def imu_worker_finished(self):
        self.imu_work = None
        pass

    def x_acc_show(self, speed):
        self.X_acc_show_lcd.display(speed)
        self.X_acc_show_lcd.show()
        pass

    def y_acc_show(self, speed):
        self.Y_acc_show_lcd.display(speed)
        self.Y_acc_show_lcd.show()
        pass

    def yaw_show(self, speed):
        self.yaw_show_lcd.display(speed)
        self.yaw_show_lcd.show()
        pass

    def mistake_message_show(self, message):
        QMessageBox.information(self, "错误", "错误原因{}".format(message))
        pass


class PortSelectionDialog_wheel(QDialog):
    port_baudrate_selected_signal = pyqtSignal(str, int)  # 修改信号，现在包括端口和波特率

    def __init__(self):
        super().__init__()
        self.setWindowFlags(QtCore.Qt.FramelessWindowHint)
        layout = QVBoxLayout(self)

        self.comboBoxPort = QComboBox()
        self.populate_ports()  # 动态加载端口
        layout.addWidget(self.comboBoxPort)

        # 波特率选择下拉菜单
        self.comboBoxBaudrate = QComboBox()
        self.comboBoxBaudrate.addItems(["115200"])  # 常用波特率
        layout.addWidget(self.comboBoxBaudrate)

        # 确认按钮
        self.confirmButton = QPushButton("确认")
        self.confirmButton.clicked.connect(self.confirm)
        layout.addWidget(self.confirmButton)

        # 关闭按钮
        self.closeButton = QPushButton("关闭")
        self.closeButton.clicked.connect(self.close)
        layout.addWidget(self.closeButton)

    def confirm(self):
        """确认按钮的事件处理器"""
        chosen_port = self.comboBoxPort.currentText()
        chosen_baudrate = int(self.comboBoxBaudrate.currentText())
        self.port_baudrate_selected_signal.emit(chosen_port, chosen_baudrate)
        self.accept()

    def populate_ports(self):
        """动态加载可用的串口到下拉菜单"""
        self.comboBoxPort.clear()  # 清空当前的选项
        ports = serial.tools.list_ports.comports()
        available_ports = [port.device for port in ports]
        self.comboBoxPort.addItems(available_ports)


class PortSelectionDialog_radar(QDialog):
    port_baudrate_selected_signal = pyqtSignal(str, int)  # 修改信号，现在包括端口和波特率

    def __init__(self):
        super().__init__()
        self.setWindowFlags(QtCore.Qt.FramelessWindowHint)
        layout = QVBoxLayout(self)

        self.comboBoxPort = QComboBox()
        self.populate_ports()  # 动态加载端口
        layout.addWidget(self.comboBoxPort)

        # 波特率选择下拉菜单
        self.comboBoxBaudrate = QComboBox()
        self.comboBoxBaudrate.addItems(["115200"])  # 常用波特率
        layout.addWidget(self.comboBoxBaudrate)

        # 确认按钮
        self.confirmButton = QPushButton("确认")
        self.confirmButton.clicked.connect(self.confirm)
        layout.addWidget(self.confirmButton)

        # 关闭按钮
        self.closeButton = QPushButton("关闭")
        self.closeButton.clicked.connect(self.close)
        layout.addWidget(self.closeButton)

    def confirm(self):
        """确认按钮的事件处理器"""
        chosen_port = self.comboBoxPort.currentText()
        chosen_baudrate = int(self.comboBoxBaudrate.currentText())
        self.port_baudrate_selected_signal.emit(chosen_port, chosen_baudrate)
        self.accept()

    def populate_ports(self):
        """动态加载可用的串口到下拉菜单"""
        self.comboBoxPort.clear()  # 清空当前的选项
        ports = serial.tools.list_ports.comports()
        available_ports = [port.device for port in ports]
        self.comboBoxPort.addItems(available_ports)


class PortSelectionDialog_imu(QDialog):
    port_baudrate_selected_signal = pyqtSignal(str, int)  # 修改信号，现在包括端口和波特率

    def __init__(self):
        super().__init__()
        self.setWindowFlags(QtCore.Qt.FramelessWindowHint)
        layout = QVBoxLayout(self)

        self.comboBoxPort = QComboBox()
        self.populate_ports()  # 动态加载端口
        layout.addWidget(self.comboBoxPort)

        # 波特率选择下拉菜单
        self.comboBoxBaudrate = QComboBox()
        self.comboBoxBaudrate.addItems(["115200"])  # 常用波特率
        layout.addWidget(self.comboBoxBaudrate)

        # 确认按钮
        self.confirmButton = QPushButton("确认")
        self.confirmButton.clicked.connect(self.confirm)
        layout.addWidget(self.confirmButton)

        # 关闭按钮
        self.closeButton = QPushButton("关闭")
        self.closeButton.clicked.connect(self.close)
        layout.addWidget(self.closeButton)

    def confirm(self):
        """确认按钮的事件处理器"""
        chosen_port = self.comboBoxPort.currentText()
        chosen_baudrate = int(self.comboBoxBaudrate.currentText())
        self.port_baudrate_selected_signal.emit(chosen_port, chosen_baudrate)
        self.accept()

    def populate_ports(self):
        """动态加载可用的串口到下拉菜单"""
        self.comboBoxPort.clear()  # 清空当前的选项
        ports = serial.tools.list_ports.comports()
        available_ports = [port.device for port in ports]
        self.comboBoxPort.addItems(available_ports)


class Worker_wheel(QThread):
    finished_signal = pyqtSignal()
    tran_speed_signal = pyqtSignal(float)
    mistake_message_transmit = pyqtSignal(str)
    data_collected_signal = pyqtSignal(list, list)

    def __init__(self, port, baudrate):
        super(Worker_wheel, self).__init__()
        self.should_run = True
        self.port = 'COM18'
        self.baudrate = baudrate
        self.speed_data = []
        self.time_stamps = []
        self.speed_data_for_avg = []

    def run(self):
        try:
            ser = serial.Serial(self.port, self.baudrate, timeout=1)
            while self.should_run:
                try:
                    data_line = ser.readline().decode('utf-8').rstrip()
                    # 拆分和解析数据
                    if "Velocity = " in data_line:
                        # 提取速度值
                        velocity_str = data_line.replace("Velocity = ", "")
                        speed = float(velocity_str)
                        print(speed)
                        self.speed_data.append(speed)
                        self.time_stamps.append(datetime.now())
                        self.tran_speed_signal.emit(speed)
                except ValueError:
                    continue
        except Exception as err_1:
            self.mistake_message_transmit.emit(str(err_1))
            self.data_collected_signal.emit(self.speed_data, self.time_stamps)
        finally:
            self.data_collected_signal.emit(self.speed_data, self.time_stamps)
            self.finished_signal.emit()


class Worker_gnss(QThread):
    finished_signal = pyqtSignal()
    mistake_message_transmit = pyqtSignal(str)
    average_speed_signal = pyqtSignal(float)
    estimated_speed_signal = pyqtSignal(float)

    data_collected_signal = pyqtSignal(list, list,list,list,list,list,list)
    message_generated = threading.Event()

    def __init__(self):
        super(Worker_gnss, self).__init__()
        self.ser = None
        self.port = 'COM15'
        self.baudrate = 115200
        self.should_run = True
        self.gnss_messages_send = None
        self.speed_data = []
        self.speed_data_for_avg = []
        self.estimated_speed_data = []
        self.estimated_speed_data_for_avg = []
        self.time_stamps = []
        self.previous_timestamp = None
        self.time_difference = 0.0
        self.radar_data = []
        self.acc_x_data = []
        self.acc_y_data = []
        self.yaw_data = []

    def run(self):
        try:
            lat, lon, alt, speed, rtk_fix_quality, time_difference, estimated_speed = 0, 0, 0, 0.0, 0, 0.0, 0.0
            initial_covariance = np.eye(3)
            Q_opt = np.diag([0.01690025, 0.01690025, 0.01690025])
            R_opt = np.diag([0.00386079942806028, 0.00386079942806028, 0.00386079942806028])
            initial_state = [0, 0, 0]
            gngga_messages = []
            estimated_speed = 0.0
            if self.baudrate is not None and self.port is not None:
                self.ser = serial.Serial(self.port, self.baudrate, timeout=1)
                self.ser.write("gnrmc 0.05\r\n".encode())
                self.ser.write("gngga 0.05\r\n".encode())
                ekf = EKF(initial_state, initial_covariance, Q_opt, R_opt)
                while self.should_run:
                    try:
                        current_timestamp = datetime.now()
                        if self.previous_timestamp:
                            self.time_difference = (current_timestamp - self.previous_timestamp).total_seconds()
                        self.previous_timestamp = current_timestamp
                        raw_message_gngga = self.ser.readline().decode('utf-8').rstrip()
                        if raw_message_gngga.startswith("$GNGGA"):
                            gngga_messages.append(raw_message_gngga)
                            fields = raw_message_gngga.split(",")
                            rtk_fix_quality = fields[6]
                            print(rtk_fix_quality)
                            if len(gngga_messages) % 20 == 0:
                                self.gnss_messages_send = "\r\n".join(gngga_messages) + "\r\n"
                                gngga_messages = []
                                self.message_generated.set()
                        raw_message_gnrmc = self.ser.readline().decode('utf-8').rstrip()
                        if raw_message_gnrmc.startswith("$GNRMC"):
                            print(raw_message_gnrmc)
                            if rtk_fix_quality == '4' or rtk_fix_quality == '5':
                                speed = extract_speed_mps(raw_message_gnrmc)
                            # speed = extract_speed_kmh(raw_message_gnrmc)
                            # print(1, speed)
                        # EKF融合速度计算
                            if speed != 0 :
                                estimated_speed = ekf.update(speed, radar_speed, acc_x, acc_y, yaw,
                                                         self.time_difference)
                                self.estimated_speed_signal.emit(estimated_speed)


                            self.radar_data.append(radar_speed)
                            self.acc_x_data.append(acc_x)
                            self.acc_y_data.append(acc_y)
                            self.yaw_data.append(yaw)

                        self.speed_data.append(speed)
                        self.speed_data_for_avg.append(speed)

                        self.estimated_speed_data.append(estimated_speed)
                        self.estimated_speed_data_for_avg.append(estimated_speed)

                        self.time_stamps.append(current_timestamp)
                        if len(self.speed_data_for_avg) == 10:
                            avg_speed = sum(self.speed_data_for_avg) / 10
                            self.average_speed_signal.emit(avg_speed)
                            self.speed_data_for_avg.clear()
                            pass

                        if len(self.estimated_speed_data_for_avg) == 10:
                            avg_estimated_speed = sum(self.estimated_speed_data_for_avg) / 10
                            self.estimated_speed_signal.emit(avg_estimated_speed)
                            self.estimated_speed_data_for_avg.clear()
                            pass
                    except Exception as err_1:
                        self.mistake_message_transmit.emit(str(err_1))
                        continue
        except Exception as err_1:
            self.mistake_message_transmit.emit(str(err_1))
            self.data_collected_signal.emit(self.speed_data,self.radar_data,self.acc_x_data,self.acc_y_data,self.yaw_data, self.estimated_speed_data, self.time_stamps)
        finally:
            self.finished_signal.emit()
            self.data_collected_signal.emit(self.speed_data, self.radar_data, self.acc_x_data, self.acc_y_data,
                                            self.yaw_data, self.estimated_speed_data, self.time_stamps)

            self.ser.close()


class Worker_Ntrip(QThread):
    finished_signal = pyqtSignal()
    mistake_message_transmit = pyqtSignal(str)

    def __init__(self, gnss_worker):
        super(Worker_Ntrip, self).__init__()
        self.port = "COM7"
        self.baudrate = 115200
        self.should_run = True
        self.gnss_worker = gnss_worker

        self.Ntrip_host = "39.107.207.235"
        self.Ntrip_port = 8002
        self.Ntrip_mountpoint = "AUTO"
        self.Ntrip_user = "qxyjjd001"
        self.Ntrip_password = "c3934ef"

    def run(self):
        try:
            ser = serial.Serial(self.port, self.baudrate, timeout=1)
            self.reconnect()
            while self.should_run:
                gnss_messages_send = self.gnss_worker.gnss_messages_send
                if gnss_messages_send == None:
                    continue
                if self.RTK_singal_OK == 1:
                    self.s.send(gnss_messages_send.encode())
                    rtcm_data = self.s.recv(1024000)
                    if rtcm_data is not None:
                        ser.write(rtcm_data)
                        ser.flush()
        except Exception as err_1:
            self.mistake_message_transmit.emit(str(err_1))
        finally:
            self.finished_signal.emit()
            ser.close()

    def reconnect(self):
        while True:
            self.s = socket.socket()
            self.s.settimeout(10)
            self.s.connect((self.Ntrip_host, self.Ntrip_port))
            self.ntrip_request = f"GET /{self.Ntrip_mountpoint} HTTP/1.0\r\n"
            self.ntrip_request += "User-Agent: NTRIP PythonClient/1.0\r\n"
            self.ntrip_request += f"Authorization: Basic {base64.b64encode((self.Ntrip_user + ':' + self.Ntrip_password).encode()).decode()}\r\n"
            self.ntrip_request += "\r\n"

            self.s.send(self.ntrip_request.encode())
            response = self.s.recv(1024)

            if b"ICY 200 OK" in response:
                self.RTK_singal_OK = 1
                print('ICY 200 OK')
                break
            else:
                print('链接失败')


class Worker_radar(QThread):
    finished_signal = pyqtSignal()
    mistake_message_transmit = pyqtSignal(str)
    data_collected_signal = pyqtSignal(list, list)
    avg_data_parsed = pyqtSignal(float)

    def __init__(self, port, baudrate):
        super(Worker_radar, self).__init__()
        self.should_run = True
        self.port = 'COM5'
        self.baudrate = baudrate
        print(self.port)
        print(self.baudrate)
        self.serial_connection = None
        self.speed_list = []
        self.time_stamps = []
        self.speed_data_for_avg = []

    def run(self):
        try:
            global radar_speed
            self.serial_connection = serial.Serial(self.port, self.baudrate, timeout=1)
            while self.should_run:
                try:
                    line = self.serial_connection.readline().decode('utf-8', errors='ignore').strip()
                    if line.startswith("R "):
                        speed = line[1:]
                        speed = float(speed)
                        radar_speed = speed
                        # 将数据存储到列表中
                        self.speed_list.append(speed)
                        self.time_stamps.append(datetime.now())
                        self.speed_data_for_avg.append(speed)
                        if len(self.speed_data_for_avg) == 5:
                            avg_speed = sum(self.speed_data_for_avg) / 5
                            self.avg_data_parsed.emit(avg_speed)
                            self.speed_data_for_avg.clear()
                except Exception as err_1:
                    self.mistake_message_transmit.emit(str(err_1))
                    continue
            self.finished_signal.emit()
        except Exception as err_1:
            self.mistake_message_transmit.emit(err_1)
            self.data_collected_signal.emit(self.speed_list, self.time_stamps)
        finally:
            self.data_collected_signal.emit(self.speed_list, self.time_stamps)
            self.finished_signal.emit()


class Worker_IMU(QThread):
    finished_signal = pyqtSignal()
    tran_x_acc_signal = pyqtSignal(float)
    tran_y_acc_signal = pyqtSignal(float)
    tran_yaw_signal = pyqtSignal(float)
    mistake_message_transmit = pyqtSignal(str)
    tran_data_collected_signal = pyqtSignal(list, list, list, list)

    def __init__(self, port, baudrate):
        super(Worker_IMU, self).__init__()
        self.should_run = True
        self.port = 'COM13'
        self.baudrate = baudrate
        self.time_stamps = []
        self.x_acc_data = []
        self.y_acc_data = []
        self.yaw_data = []
        self.x_acc_for_avg = []
        self.y_acc_for_avg = []
        self.yaw_for_avg = []
        self.x_acc_values = []
        self.y_acc_values = []
        self.yaw_values = []

    def run(self):
        try:
            global acc_x, acc_y, yaw
            self.ser = serial.Serial(self.port, self.baudrate, timeout=1)
            while self.should_run:
                try:
                    data_line = self.ser.readline().decode('utf-8').rstrip()
                    # 拆分和解析数据
                    if data_line.startswith("acc_x:"):
                        parts = data_line.split()
                        self.time_stamps.append(datetime.now())

                        acc_x = float(parts[0].split(":")[1])
                        self.x_acc_data.append(acc_x)
                        acc_y = float(parts[1].split(":")[1])
                        self.y_acc_data.append(acc_y)

                        self.x_acc_values.append(acc_x)
                        self.y_acc_values.append(acc_y)
                        pass
                    elif data_line.startswith("angle_y:"):
                        yaw = float(data_line.split(":")[1])
                        self.yaw_data.append(yaw)
                        self.yaw_values.append(yaw)
                        pass

                    if len(self.x_acc_values) >= 10:
                        self.x_acc_for_avg = sum(self.x_acc_values) / len(self.x_acc_values)
                        self.y_acc_for_avg = sum(self.y_acc_values) / len(self.y_acc_values)
                        self.yaw_for_avg = sum(self.yaw_values) / len(self.yaw_values)

                        self.x_acc_values.clear()
                        self.y_acc_values.clear()
                        self.yaw_values.clear()

                        self.tran_x_acc_signal.emit(self.x_acc_for_avg)
                        self.tran_y_acc_signal.emit(self.y_acc_for_avg)
                        self.tran_yaw_signal.emit(self.yaw_for_avg)
                        pass
                    pass
                except ValueError:
                    continue
                    pass
        except Exception as err_1:
            self.mistake_message_transmit.emit(str(err_1))
            self.tran_data_collected_signal.emit(self.x_acc_data, self.y_acc_data, self.yaw_data, self.time_stamps)
            pass
        finally:
            self.ser.close()
            self.tran_data_collected_signal.emit(self.x_acc_data, self.y_acc_data, self.yaw_data, self.time_stamps)
            self.finished_signal.emit()
            pass


if __name__ == '__main__':
    try:
        QApplication.setAttribute(Qt.AA_EnableHighDpiScaling)
        app = QApplication(sys.argv)
        stacked_widget = QStackedWidget()
        Welcome_Window = WelcomeWindow(stacked_widget)
        information_Show_Window = information_show_Window(stacked_widget)
        stacked_widget.addWidget(Welcome_Window)
        stacked_widget.addWidget(information_Show_Window)
        stacked_widget.setCurrentIndex(0)
        main_window = QMainWindow()  # 创建一个主窗口
        main_window.setWindowTitle("果园作业机械研究团队")  # 设置标题
        main_window.setCentralWidget(stacked_widget)  # 将stacked_widget设置为主窗口的中心小部件
        main_window.show()
        sys.exit(app.exec_())
    except Exception as err:
        print(err)
